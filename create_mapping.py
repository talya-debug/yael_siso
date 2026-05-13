import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Billing Mapping'

header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color='091426', end_color='091426', fill_type='solid')
yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

headers = ['Project', 'Billing Milestone', 'Amount', 'Linked Scope Phase', 'Trigger', 'Chloe OK?']
ws.append(headers)
for col in range(1, 7):
    cell = ws.cell(row=1, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 45
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 30
ws.column_dimensions['E'].width = 20
ws.column_dimensions['F'].width = 12

data = [
    ['Feivel 5', 'After Closing the working plans', 25500, 'Working Plans - Initial', 'Phase end', '', False],
    ['Feivel 5', 'Upon project completion', 8500, 'Project Completion', 'Phase end', '', False],
    ['Moma', 'Project Completion', 6500, 'Project Completion', 'Phase end', '', False],
    ['Shaul Hamelech 4', 'After completing the working plans', 23000, 'Working Plans - Initial', 'Phase end', '', False],
    ['Shaul Hamelech 4', 'At the start of construction work', 11500, 'Execution', '??? start or end?', '', True],
    ['Shaul Hamelech 4', 'After Carpentry plans', 23000, 'Working Plans - Advanced', 'Phase end', '', False],
    ['Shaul Hamelech 4', 'Upon project completion', 11500, 'Project Completion', 'Phase end', '', False],
    ['Nachmani 15', 'After completing the working plans', 25650, 'Working Plans - Initial', 'Phase end', '', False],
    ['Nachmani 15', 'During the supervision stage', 25650, 'Execution', '??? start or end?', '', True],
    ['Nachmani 15', 'Upon project completion', 12825, 'Project Completion', 'Phase end', '', False],
    ['Miryam Haheshmonait', 'After selecting kitchen and furniture', 32500, 'Basic Material Selection', 'Phase end', '', False],
    ['Miryam Haheshmonait', 'Upon project completion', 13000, 'Project Completion', 'Phase end', '', False],
    ['Mosinzon 5', 'Upon completion of the work', 10000, 'Project Completion', 'Phase end', '', False],
    ['Bilu 4', 'Receipt of occupancy approval', 48750, 'Project Completion', 'Phase end', '', False],
    ['Bilu 4', 'Additional Payment - Plan Changes', 0, 'Extra (Yael decides)', '???', '', True],
    ['Geula', 'After completing the working plans', 33000, 'Working Plans - Initial', 'Phase end', '', False],
    ['Geula', 'During the supervision stage', 11000, 'Execution', '??? start or end?', '', True],
    ['Geula', 'Upon project completion', 11000, 'Project Completion', 'Phase end', '', False],
]

for row_data in data:
    highlight = row_data.pop()
    ws.append(row_data)
    if highlight:
        row_idx = ws.max_row
        for col in range(1, 7):
            ws.cell(row=row_idx, column=col).fill = yellow_fill

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center' if cell.column > 2 else 'left')

wb.save('billing_mapping_for_chloe.xlsx')
print('Done')
